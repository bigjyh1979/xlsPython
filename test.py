#-*- coding: utf-8 -*-
'''
@author: birdlin
'''
import openpyxl
capacity = 0
Wh = 0

wb = openpyxl.load_workbook('rest+discharge.xlsx', read_only=False, )
print (wb.get_sheet_names())

ws = wb.get_sheet_by_name ('sheet1')
print (ws['B2'].value)
print (type(ws['B2'].value))
print (ws.max_row)
print (ws.max_column)

for row in ws.rows:
    #for cell in row:
    print (row)
    print ()


#print (ws.cell(column=2, row=2).value)


'''
with open ('rest+discharge.txt', 'r', encoding='utf-8') as fin:
    with open ('newtestCSV.txt', 'w', encoding='utf-8') as fout :
        csvreader = csv.reader(fin, delimiter=',')
        csvwriter = csv.writer(fout, delimiter=',')
        header = next(csvreader)
        header.append('mAh')                        ## V,I,P,R,T,mAH,WH,mAh
        header.append('Wh')                         ## V,I,P,R,T,mAH,WH,mAh,Wh
        csvwriter.writerow(header)
        for row in csvreader:
            row[0]= round(float(row[0]), 3)         ## Voltage
            row[1]= round(float(row[1]), 3)         ## Current
            row[2]= round(float(row[2]), 3)         ## I,P,R,T,mAH,WH
            row[3]= round(float(row[3]), 3)         ## Resistance
            row[4]= round(float(row[4]), 3)         ## Temperature
            row[5]= round(float(row[5]), 3)         ## mAh - LM
            row[6]= round(float(row[6]), 3)         ## Wh - LM
            capacity += row[1] * 1000 * 10/3600     ## mAh
            Wh += row[1] * row[0] * 10/3600         ## Wh
            row.append(round(capacity, 3))          ## Extend Row
            row.append(round(Wh, 3))                ## Extend Row
            csvwriter.writerow(row)
'''
